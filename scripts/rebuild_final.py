# -*- coding: utf-8 -*-
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime, timedelta
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')

# ============================================================
# 1. Read & merge data (same logic as before)
# ============================================================
with open(os.path.join(DATA_DIR, 'sui_song_list_complete.json'), 'r', encoding='utf-8') as f:
    web_songs = json.load(f)

web_data = {}
for item in web_songs:
    name = item['song_name'].strip()
    if not name:
        name = '(歌曲名缺失-源数据问题)'
    dates = set()
    if item.get('date_list'):
        for d in item['date_list'].split('\uff0c'):
            d = d.strip()
            if d:
                try:
                    dt = datetime.strptime(d, '%Y/%m/%d')
                    dates.add(dt)
                except:
                    pass
    web_data[name] = {
        'dates': dates,
        'artist': item.get('artist', '') or '',
        'lang': item.get('language', ''),
        'translated': item.get('song_translated_name', '') or '',
        'remarks': item.get('remarks', '') or '',
        'bvids': item.get('BVID', '') or '',
        'song_count': item.get('song_count', 0),
    }

src = openpyxl.load_workbook(os.path.join(DATA_DIR, '岁己数据统计.xlsx'), read_only=True)
ws_src = src[src.sheetnames[1]]
local_data = {}
current_date = None
for row in ws_src.iter_rows(min_row=2, values_only=True):
    date_val, song_val = row[0], row[1]
    if date_val is not None:
        current_date = date_val
    if song_val is not None:
        name = str(song_val).strip()
        if not name:
            continue
        dt = None
        try:
            if isinstance(current_date, (int, float)):
                dt = datetime(1899, 12, 30) + timedelta(days=int(current_date))
            elif isinstance(current_date, datetime):
                dt = current_date
        except:
            pass
        local_data.setdefault(name, set())
        if dt:
            local_data[name].add(dt)

lang_map = {'华语': '中文', '日语': '日语', '英语': '英文', '韩语': '韩语', '中文': '中文', '英文': '英文'}

def detect_lang(s):
    has_jp = bool(re.search(r'[\u3040-\u309f\u30a0-\u30ff]', s))
    has_cn = bool(re.search(r'[\u4e00-\u9fff]', s))
    has_en = bool(re.search(r'[a-zA-Z]', s))
    has_kr = bool(re.search(r'[\uac00-\ud7af\u1100-\u11ff]', s))
    if has_jp: return '日语'
    if has_kr: return '韩语'
    if has_cn and not has_en: return '中文'
    if has_en and not has_cn: return '英文'
    if has_cn and has_en: return '中英混合'
    return '其他'

LOCAL_CUTOFF = datetime(2024, 6, 1)
merged = {}
for name in set(web_data.keys()) | set(local_data.keys()):
    w = web_data.get(name, {})
    l = local_data.get(name, set())
    web_dates = w.get('dates', set()) if w else set()
    if l:
        # Local data exists: use local dates + web dates only before 2024-06
        all_dates = l | {d for d in web_dates if d < LOCAL_CUTOFF}
    else:
        # No local data: use all web dates
        all_dates = web_dates
    # Use date count, fallback to song_count from JSON if dates are empty
    date_count = len(all_dates)
    web_song_count = w.get('song_count', 0) if w else 0
    final_count = max(date_count, web_song_count)
    artist = (w.get('artist', '') if w else '') or ''
    lang_raw = (w.get('lang', '') if w else '')
    lang = lang_map.get(lang_raw, '') if lang_raw else detect_lang(name)
    first_date = min(all_dates) if all_dates else None
    last_date = max(all_dates) if all_dates else None
    merged[name] = {
        'song': name, 'count': final_count, 'artist': artist,
        'lang': lang, 'translated': (w.get('translated', '') if w else '') or '',
        'remarks': (w.get('remarks', '') if w else '') or '',
        'first': first_date, 'last': last_date,
    }

songs = sorted(merged.values(), key=lambda x: (-x['count'], x['song']))

# ── Date corrections (user-confirmed erroneous dates) ──
date_corrections = {
    'first love': {datetime(2024, 7, 28)},
    '雪峰 ～yukimine～': {datetime(2024, 7, 28)},
}
for s in songs:
    if s['song'] in date_corrections:
        # We can't modify dates directly since they're already aggregated,
        # so we rebuild from source
        pass

# Rebuild corrected songs
corrected = []
for s in songs:
    name = s['song']
    w = web_data.get(name, {})
    l = local_data.get(name, set())
    web_dates = w.get('dates', set()) if w else set()
    if l:
        all_dates = l | {d for d in web_dates if d < LOCAL_CUTOFF}
    else:
        all_dates = web_dates

    # Apply corrections
    if name in date_corrections:
        all_dates -= date_corrections[name]

    date_count = len(all_dates)
    web_song_count = w.get('song_count', 0) if w else 0
    final_count = max(date_count, web_song_count)
    if name in date_corrections and date_count < web_song_count:
        final_count = date_count  # Don't use web fallback for corrected songs

    first_date = min(all_dates) if all_dates else None
    last_date = max(all_dates) if all_dates else None
    s['count'] = final_count
    s['first'] = first_date
    s['last'] = last_date

# Re-sort after corrections
songs.sort(key=lambda x: (-x['count'], x['song']))

# ── Language overrides (user-confirmed corrections) ──
lang_overrides = {
    '交织together': '中文',      # has English in name but is Chinese song
    '问（DJ版）': '中文',        # DJ is just a version tag
    'Alice in 冷凍庫': '日语',   # Japanese song (Orangestar)
    '8.32': '日语',              # Japanese Vocaloid song (*Luna)
}

# ── Artist overrides (user-confirmed corrections) ──
artist_overrides = {
    '8.32': '*Luna',
    'Alice in 冷凍庫': 'Orangestar',
    '懐中道标': 'やなぎなぎ',
    "Don't cry Don't cry": '魏如萱',
}

for s in songs:
    if s['song'] in lang_overrides:
        s['lang'] = lang_overrides[s['song']]
    if s['song'] in artist_overrides:
        s['artist'] = artist_overrides[s['song']]

# Artist mapping from web research (fills gaps where source data has no artist)
artist_map = {
    '爱情讯息': '郭静', '爱错': '王力宏', '睡吧，睡吧，我亲爱的宝贝': '舒伯特（传统摇篮曲）',
    '蝴蝶': '陶喆', '下等马': '陈壹千', '交织together': '泠鸢yousa', '又三郎': 'ヨルシカ',
    '夏霞': 'aiko', '夕日坂': 'Doriko', '小幸运': '田馥甄', '尚好的青春': '孙燕姿',
    '无赖': '郑中基', '虫儿飞': '郑伊健（《风云》插曲）', '鸟之诗': 'Lia（《AIR》OP）',
    '一人行者': '蔡健雅', '上海一九四三': '周杰伦', '两个恰恰好': '旺福',
    '亲爱的，那不是爱情': '张韶涵', '你不知道的事': '王力宏', '初恋': '莫文蔚',
    '别那么骄傲': '金海心', '加油鸭': '小旭音乐', '化身孤岛的鲸': '周深',
    '原来你什么都不要': '张惠妹', '友谊地久天长': '传统苏格兰民歌', '友谊天长地久': '传统苏格兰民歌',
    '回留': '蔡健雅', '地狱先生': '小缘', '夜宴风波': '音阙诗听/赵方婧',
    '大东北我的家乡': '二手玫瑰', '天使': '五月天', '如果没有你': '莫文蔚',
    '妄想感傷代償連盟': 'DECO*27', '小宇': '张震岳', '屑屑': 'ChiliChill',
    '干物女': '封茗囧菌', '心愿便利贴': '元若蓝', '心拍数#0822': '蝶々P',
    '怎么了': '周兴哲', '思想犯': '陈壹千', '恋爱语音导航': '洛天依',
    '悠哉日常': 'のんのんびより', '明天你好': '牛奶咖啡',
    '易燃易爆炸': '陈粒', '星星和雨的夜': '银临', '星间飞行': '中島愛',
    '月出': '双笙', '有形的翅膀': '张韶涵', '欧若拉': '张韶涵',
    '深海少女': 'ゆうゆ（初音ミク）', '火葬场之歌': 'ilem（洛天依）', '爆刘继芬': '茶理理',
    '爱你但说不出口': '王贰浪', '狐狸精': '徐良', '眼镜的葬礼': '小旭音乐',
    '社畜烧酒': '泠鸢yousa', '神的随波逐流': 'そらる/まふまふ', '私奔到月球': '五月天/陈绮贞',
    '童话': '光良', '第几个一百天': '林俊杰', '给我一个理由忘记': 'A-Lin',
    '给我一首歌的时间': '周杰伦', '绮凝盏': '双笙', '群青日和': '椎名林檎',
    '老娘驾到': '阿肆', '花,太阳,彩虹,你': '唐磊', '茉莉花': '传统民歌',
    '莹梦': '双笙', '藍二乗': 'aiko', '说唱脸谱': '谢津', '败家娘们儿': '大庆小芳',
    '走在冷风中': '刘思涵', '起风了': '买辣椒也用券', '送别': '李叔同（传统）',
    '遗失的美好': '张韶涵', '采茶纪': '双笙', '阳光下的星星': '金海心',
    '隐形的翅膀': '张韶涵', '雪绒花': '传统（《音乐之声》）', '页角情书': '双笙',
    '问（DJ版）': '李宗盛',
    'Rainbow Girl': '40mP', 'あの夏が飽和する。': 'カンザキイオリ',
    'グリグリメガネと月光蟲': 'ハチ（米津玄师）', 'メルト': 'ryo/supercell（初音ミク）',
    '怪獣の花唄': 'Vaundy', '言って。': 'ヨルシカ', '都落ち': 'ヨルシカ',
    'あおぞら': '40mP（初音ミク）', 'シュガーソングとビターステップ': 'UNISON SQUARE GARDEN',
    'ビードロ模様': 'やなぎなぎ', '命に嫌われている。': 'カンザキイオリ',
    '悪魔の子': 'ヒグチアイ', '私じゃなかったんだね': 'Aimer',
    'あなたがいた森': '樹海', 'さよならの夏': '手嶌葵', 'とびら開けて': '神前暁',
    'とりのこしてィ': 'まふまふ', 'ひまわりの約束': '秦基博', 'アイディスマイル': 'tuki.',
    'エウテルペ': 'EGOIST', 'キセキ～未来へ～': 'HoneyWorks', 'キミの記憶': 'Kalafina',
    'クリームソーダとシャンデリア': 'ハチ（米津玄师）', 'コールボーイ': 'syudou',
    'サマータイムレコード': 'じん（初音ミク）', 'センチメンタルな愛慕心': 'ハチ（米津玄师）',
    'センチメートル': 'the peggies', 'ダイアモンド クレバス': "May'n/中島愛",
    'トリノコシティ': '164（初音ミク）', 'ハイドアンド・シーク': 'バルーン',
    'メリュー': 'ハチ（米津玄师）', 'メリーメリ': 'バルーン', '刹那プラス': 'バルーン',
    '割れたリンゴ': '渡辺麻友', '君との明日': '西野カナ', '嘘つき': 'バルーン',
    '好き！雪！本気マジック': 'Mitchie M（初音ミク）', '心做し': '蝶々P（GUMI）',
    '続・へたくそユートピア政策': 'バルーン', '行かないで': '中森明菜',
    '貴方の恋人になりたい': 'まふまふ', '運命のルーレット廻して': 'ZARD',
    '雪は何色': 'LiSA', '風に薫る夏の記憶': 'コブクロ', '風を食む': 'ヨルシカ',
    'For フルーツバスケット': '岡崎律子',
    'シリョクケンサ（视力检查）': '40mP（初音ミク）',
    'シンクロサイクロトロン・スピリチュアライザー': 'ハチ（米津玄师）',
    'シンクロサイクロトロン・スピリチュアライザー。': 'ハチ（米津玄师）',
    '歩いても 歩いても': 'くるり', '莉回る空うさぎ': 'ハチ（米津玄师）',
    '给远在天边的你': '中島愛',
    'Letting Go': '蔡健雅', 'Fly Me to the Moon': 'Frank Sinatra', 'me me she': 'Bruno Mars',
    'From The Start': 'Laufey', 'Ghost of a smile': 'EGOIST', 'Plastic Love': '竹内まりや',
    'City of Stars': 'Ryan Gosling/Emma Stone', 'DAYBREAK FRONTLINE': 'Linked Horizon',
    'Just The Two Of Us': 'Grover Washington Jr.', 'One Last Kiss': '宇多田ヒカル',
    'Remember me': 'Various（《Coco》）', 'WHITE ALBUM': '平野綾',
    'from Y to Y': 'Jimmy Eat World', 'rain stops, good-bye': 'The Script',
    'After 17': '陈绮贞', 'Again': 'YUI', 'Alice in 冷凍庫': 'ChiliChill',
    'Departures': 'EGOIST', "Don't Look Back in Anger": 'Oasis',
    "Don't cry Don't cry": 'ChiliChill', 'Drop Pop Candy': 'Giga/Reol',
    'First Love': '宇多田ヒカル', 'Forever Young': 'Bob Dylan',
    'I Love You 3000': 'Stephanie Poetri',
    'I Really Want to Stay at Your House': 'Rosa Walton/Hallie Coggins',
    'Kiss Me': 'Sixpence None the Richer', 'LOVE 2000': '安室奈美恵',
    'Let It Be': 'The Beatles', 'Letter Song': 'doriko（初音ミク）', 'Lover': 'Taylor Swift',
    'No title': 'Reol', 'One More Light': 'Linkin Park', 'Say So': 'Doja Cat',
    'Someone Like You': 'Adele', 'Stay With Me': 'Sam Smith',
    'Try Everything': 'Shakira', 'Upupu': 'Aimer',
    'Virtual To LIVE': 'にじさんじ', 'WHITEEALBUM': '平野綾',
    'hacking to the gate': 'いとうかなこ', 'lemon': '米津玄師',
    'one more time one more chance': '山崎まさよし', 'stand by me': 'Ben E. King',
    'tune the rainbow': '坂本真綾', '8.32': 'ChiliChill',
    'rain stops,good-bye': 'The Script', 'ひまわりの约束': '秦基博',
    # Jul/Aug 2024 new songs (not in web data)
    'God knows': '神前暁（《涼宮ハルヒの憂鬱》）',
    'Someone You Loved': 'Lewis Capaldi',
    '修炼爱情': '林俊杰',
    '王妃': '萧敬腾',
}

def normalize(s):
    s = s.strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('\u2011', '-').replace('\u2010', '-').replace('\u2013', '-').replace('\u2014', '-')
    return s

norm_map = {normalize(k): v for k, v in artist_map.items()}

def find_artist(name):
    if name in artist_map:
        return artist_map[name]
    n = normalize(name)
    if n in norm_map:
        return norm_map[n]
    return None

# Apply artist_map to fill gaps
for s in songs:
    if not s['artist']:
        a = find_artist(s['song'])
        if a:
            s['artist'] = a

print(f'Total: {len(songs)} songs, {sum(s["count"] for s in songs)} performances')

# ============================================================
# 2. Modern design system
# ============================================================
# Color palette - soft, modern, internet-trendy
C_PRIMARY = '2D2D3F'       # Deep charcoal for headers
C_ACCENT = '6C5CE7'        # Soft purple accent
C_ACCENT_LIGHT = 'A29BFE'  # Light lavender
C_BG_STRIPE = 'F8F7FF'     # Very subtle lavender stripe
C_BG_WHITE = 'FFFFFF'       # White
C_TIER_HIGH = 'E17055'     # Warm coral for frequent
C_TIER_MED = '6C5CE7'      # Purple for occasional
C_TIER_LOW = 'B2BEC3'      # Muted gray for rare
C_TEXT = '2D3436'           # Near-black text
C_TEXT_LIGHT = '636E72'     # Gray text
C_BORDER = 'DFE6E9'        # Light gray border
C_GOLD = 'FDCB6E'          # Soft gold for top highlights
C_HEADER_TEXT = 'FFFFFF'   # White header text

font_title = Font(name='Microsoft YaHei', bold=True, size=16, color=C_PRIMARY)
font_subtitle = Font(name='Microsoft YaHei', size=10, color=C_TEXT_LIGHT)
font_header = Font(name='Microsoft YaHei', bold=True, size=10.5, color=C_HEADER_TEXT)
font_data = Font(name='Microsoft YaHei', size=10, color=C_TEXT)
font_data_light = Font(name='Microsoft YaHei', size=9, color=C_TEXT_LIGHT)
font_tier_high = Font(name='Microsoft YaHei', size=10, color=C_TIER_HIGH, bold=True)
font_tier_med = Font(name='Microsoft YaHei', size=10, color=C_TIER_MED)
font_tier_low = Font(name='Microsoft YaHei', size=9, color=C_TIER_LOW)
font_section = Font(name='Microsoft YaHei', bold=True, size=12, color=C_ACCENT)
font_count = Font(name='Microsoft YaHei', size=10, color=C_ACCENT, bold=True)

fill_header = PatternFill('solid', fgColor=C_PRIMARY)
fill_stripe = PatternFill('solid', fgColor=C_BG_STRIPE)
fill_white = PatternFill('solid', fgColor=C_BG_WHITE)
fill_gold = PatternFill('solid', fgColor='FFF9E6')
fill_section = PatternFill('solid', fgColor='F0EDFF')
fill_sub_header = PatternFill('solid', fgColor='EDEEF2')

border_thin = Border(
    left=Side(style='thin', color=C_BORDER),
    right=Side(style='thin', color=C_BORDER),
    top=Side(style='thin', color=C_BORDER),
    bottom=Side(style='thin', color=C_BORDER))
border_none = Border()
border_bottom_only = Border(bottom=Side(style='thin', color=C_BORDER))

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    return cell

def write_header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        set_cell(ws, row, c, h, font_header, fill_header, align_center, border_thin)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def write_data_row(ws, row, values, is_stripe=False, fonts=None, aligns=None):
    bg = fill_stripe if is_stripe else fill_white
    for c, v in enumerate(values, 1):
        f = fonts[c-1] if fonts and c-1 < len(fonts) else font_data
        a = aligns[c-1] if aligns and c-1 < len(aligns) else align_center
        set_cell(ws, row, c, v, f, bg, a, border_thin)

# ============================================================
# 3. Build workbook
# ============================================================
wb = openpyxl.Workbook()

# --- Sheet 1: 完整歌单 ---
ws1 = wb.active
ws1.title = '完整歌单'
ws1.sheet_properties.tabColor = C_ACCENT

# Title area
ws1.merge_cells('A1:I1')
set_cell(ws1, 1, 1, '岁己SUI 完整歌单', font_title, None, Alignment(horizontal='left', vertical='center'))
ws1.row_dimensions[1].height = 36

total_perf = sum(s['count'] for s in songs)
freq_n = sum(1 for s in songs if s['count'] >= 5)
occ_n = sum(1 for s in songs if 2 <= s['count'] <= 4)
once_n = sum(1 for s in songs if s['count'] == 1)

ws1.merge_cells('A2:I2')
stats = f'{len(songs)} 首歌  ·  {total_perf} 次演唱  ·  常唱 {freq_n}  ·  偶尔 {occ_n}  ·  仅1次 {once_n}  ·  2022.09 – 2026.06'
set_cell(ws1, 2, 1, stats, font_subtitle, None, Alignment(horizontal='left', vertical='center'))
ws1.row_dimensions[2].height = 22

# Spacer
ws1.row_dimensions[3].height = 8

# Header
headers1 = ['#', '歌曲名', '译名', '原唱', '语言', '次数', '频率', '首次', '最近']
widths1 = [5, 30, 16, 20, 7, 6, 6, 11, 11]
write_header_row(ws1, 4, headers1, widths1)
ws1.row_dimensions[4].height = 24

# Data
for i, s in enumerate(songs, 1):
    row = i + 4
    is_stripe = i % 2 == 0
    tier = '常唱' if s['count'] >= 5 else ('偶尔' if s['count'] >= 2 else '稀有')
    tier_font = font_tier_high if s['count'] >= 5 else (font_tier_med if s['count'] >= 2 else font_tier_low)

    vals = [
        i,
        s['song'],
        s['translated'],
        s['artist'],
        s['lang'],
        s['count'],
        tier,
        s['first'].strftime('%Y-%m-%d') if s['first'] else '',
        s['last'].strftime('%Y-%m-%d') if s['last'] else '',
    ]
    fonts = [font_data_light, font_data, font_data_light, font_data, font_data, font_count, tier_font, font_data_light, font_data_light]
    aligns = [align_center, align_left, align_left, align_left, align_center, align_center, align_center, align_center, align_center]
    write_data_row(ws1, row, vals, is_stripe, fonts, aligns)

# Freeze panes
ws1.freeze_panes = 'A5'
ws1.auto_filter.ref = f'A4:I{4 + len(songs)}'

# --- Sheet 2: 常唱歌曲 ---
ws2 = wb.create_sheet('常唱歌曲')
ws2.sheet_properties.tabColor = C_TIER_HIGH

frequent = [s for s in songs if s['count'] >= 5]

ws2.merge_cells('A1:H1')
set_cell(ws2, 1, 1, '岁己SUI 常唱歌曲', font_title, None, Alignment(horizontal='left', vertical='center'))
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:H2')
set_cell(ws2, 2, 1, f'{len(frequent)} 首歌曲，演唱 5 次及以上', font_subtitle, None, Alignment(horizontal='left', vertical='center'))
ws2.row_dimensions[2].height = 22
ws2.row_dimensions[3].height = 8

headers2 = ['#', '歌曲名', '原唱', '语言', '次数', '首次', '最近', '备注']
widths2 = [5, 32, 20, 7, 6, 11, 11, 28]
write_header_row(ws2, 4, headers2, widths2)
ws2.row_dimensions[4].height = 24

for i, s in enumerate(frequent, 1):
    row = i + 4
    is_stripe = i % 2 == 0
    is_top10 = i <= 10
    bg = fill_gold if is_top10 else (fill_stripe if is_stripe else fill_white)

    vals = [
        i, s['song'], s['artist'], s['lang'], s['count'],
        s['first'].strftime('%Y-%m-%d') if s['first'] else '',
        s['last'].strftime('%Y-%m-%d') if s['last'] else '',
        s['remarks'],
    ]
    fonts_row = [font_data_light, font_data, font_data, font_data, font_count, font_data_light, font_data_light, font_data_light]
    aligns_row = [align_center, align_left, align_left, align_center, align_center, align_center, align_center, align_left]
    for c, v in enumerate(vals, 1):
        set_cell(ws2, row, c, v, fonts_row[c-1], bg, aligns_row[c-1], border_thin)

ws2.freeze_panes = 'A5'

# --- Sheet 3: 按语言分类 ---
ws3 = wb.create_sheet('按语言分类')
ws3.sheet_properties.tabColor = '00B894'

ws3.merge_cells('A1:F1')
set_cell(ws3, 1, 1, '岁己SUI 歌单 · 按语言分类', font_title, None, Alignment(horizontal='left', vertical='center'))
ws3.row_dimensions[1].height = 36
ws3.row_dimensions[2].height = 8

lang_groups = {}
for s in songs:
    lang_groups.setdefault(s['lang'], []).append(s)

current_row = 3
widths3 = [5, 32, 20, 7, 11, 11]
for c, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

lang_icons = {'中文': '华', '日语': '日', '英文': '英', '中英混合': '混', '韩语': '韩', '其他': '他'}

for lang in ['中文', '日语', '英文', '中英混合', '韩语', '其他']:
    if lang not in lang_groups:
        continue
    group = lang_groups[lang]
    ws3.merge_cells(f'A{current_row}:F{current_row}')
    icon = lang_icons.get(lang, '')
    set_cell(ws3, current_row, 1, f'{icon}  {lang}  ({len(group)}首)', font_section, fill_section, Alignment(vertical='center'))
    ws3.row_dimensions[current_row].height = 28
    current_row += 1

    sub_h = ['#', '歌曲名', '原唱', '次数', '首次', '最近']
    for c, h in enumerate(sub_h, 1):
        set_cell(ws3, current_row, c, h, Font(name='Microsoft YaHei', bold=True, size=9.5, color=C_TEXT_LIGHT), fill_sub_header, align_center, border_thin)
    ws3.row_dimensions[current_row].height = 22
    current_row += 1

    for i, s in enumerate(group, 1):
        is_stripe = i % 2 == 0
        bg = fill_stripe if is_stripe else fill_white
        vals = [
            i, s['song'], s['artist'], s['count'],
            s['first'].strftime('%Y-%m-%d') if s['first'] else '',
            s['last'].strftime('%Y-%m-%d') if s['last'] else '',
        ]
        fonts_row = [font_data_light, font_data, font_data, font_count, font_data_light, font_data_light]
        aligns_row = [align_center, align_left, align_left, align_center, align_center, align_center]
        for c, v in enumerate(vals, 1):
            set_cell(ws3, current_row, c, v, fonts_row[c-1], bg, aligns_row[c-1], border_thin)
        current_row += 1
    current_row += 1

# --- Sheet 4: 按原唱分类 ---
ws4 = wb.create_sheet('按原唱分类')
ws4.sheet_properties.tabColor = '0984E3'

ws4.merge_cells('A1:E1')
set_cell(ws4, 1, 1, '岁己SUI 歌单 · 按原唱分类', font_title, None, Alignment(horizontal='left', vertical='center'))
ws4.row_dimensions[1].height = 36
ws4.row_dimensions[2].height = 8

artist_groups = {}
for s in songs:
    a = s['artist'] if s['artist'] else '(未知)'
    artist_groups.setdefault(a, []).append(s)

sorted_artists = sorted(artist_groups.keys(), key=lambda a: -len(artist_groups[a]))

current_row = 3
widths4 = [5, 32, 7, 11, 11]
for c, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

for artist in sorted_artists:
    group = artist_groups[artist]
    total_c = sum(s['count'] for s in group)
    ws4.merge_cells(f'A{current_row}:E{current_row}')
    set_cell(ws4, current_row, 1, f'{artist}  ({len(group)}首 · {total_c}次)', font_section, fill_section, Alignment(vertical='center'))
    ws4.row_dimensions[current_row].height = 26
    current_row += 1

    sub_h = ['#', '歌曲名', '次数', '首次', '最近']
    for c, h in enumerate(sub_h, 1):
        set_cell(ws4, current_row, c, h, Font(name='Microsoft YaHei', bold=True, size=9.5, color=C_TEXT_LIGHT), fill_sub_header, align_center, border_thin)
    current_row += 1

    sorted_group = sorted(group, key=lambda x: -x['count'])
    for i, s in enumerate(sorted_group, 1):
        is_stripe = i % 2 == 0
        bg = fill_stripe if is_stripe else fill_white
        vals = [
            i, s['song'], s['count'],
            s['first'].strftime('%Y-%m-%d') if s['first'] else '',
            s['last'].strftime('%Y-%m-%d') if s['last'] else '',
        ]
        fonts_row = [font_data_light, font_data, font_count, font_data_light, font_data_light]
        aligns_row = [align_center, align_left, align_center, align_center, align_center]
        for c, v in enumerate(vals, 1):
            set_cell(ws4, current_row, c, v, fonts_row[c-1], bg, aligns_row[c-1], border_thin)
        current_row += 1
    current_row += 1

# ============================================================
# 4. Save
# ============================================================
out = os.path.join(PROJECT_ROOT, 'outputs', '岁己SUI完整歌单.xlsx')
wb.save(out)
print(f'Saved: {out}')
